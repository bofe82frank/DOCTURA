"""
Excel output engine for DocTura Desktop.

Supports multiple layout modes and metadata sheet generation.
"""

from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from docutura.core.models import (
    DocumentMetadata,
    ExcelLayoutMode,
    ExtractedTable,
    ValidationReport,
)
from docutura.core.themes import Theme


class ExcelWriter:
    """Writes extracted tables to Excel with configurable layouts."""

    def __init__(
        self,
        layout_mode: ExcelLayoutMode = ExcelLayoutMode.SEPARATE_SHEETS,
        add_borders: bool = True,
        freeze_headers: bool = True,
        theme: Optional[Theme] = None,
    ):
        """
        Initialize Excel writer.

        Args:
            layout_mode: Layout mode for logical tables
            add_borders: Add borders to cells
            freeze_headers: Freeze header row
            theme: Theme for styling
        """
        self.layout_mode = layout_mode
        self.add_borders = add_borders
        self.freeze_headers = freeze_headers
        self.theme = theme

    def write_to_excel(
        self,
        output_path: Path,
        page_tables: List[ExtractedTable],
        logical_tables: List[ExtractedTable],
        metadata: Optional[DocumentMetadata] = None,
        validation_report: Optional[ValidationReport] = None,
    ) -> None:
        """
        Write tables to Excel file.

        Args:
            output_path: Output file path
            page_tables: Page-preserved tables
            logical_tables: Logical tables
            metadata: Document metadata
            validation_report: Validation report
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Write metadata sheet first (if enabled)
        if metadata:
            self._write_metadata_sheet(wb, metadata)

        # Write page-preserved tables
        for idx, table in enumerate(page_tables, start=1):
            sheet_name = f"Page_{idx:02d}"
            self._write_table_to_sheet(wb, table, sheet_name)

        # Write logical tables based on layout mode
        if logical_tables:
            if self.layout_mode == ExcelLayoutMode.SEPARATE_SHEETS:
                self._write_logical_tables_separate(wb, logical_tables)
            elif self.layout_mode == ExcelLayoutMode.SINGLE_SHEET_VERTICAL:
                self._write_logical_tables_single_vertical(wb, logical_tables)
            elif self.layout_mode == ExcelLayoutMode.SINGLE_SHEET_HORIZONTAL:
                self._write_logical_tables_single_horizontal(wb, logical_tables)

        # Write validation sheet (if validation report exists)
        if validation_report:
            self._write_validation_sheet(wb, validation_report)

        # Save workbook
        wb.save(output_path)

    def _write_metadata_sheet(
        self, wb: Workbook, metadata: DocumentMetadata
    ) -> None:
        """Write Document_Metadata sheet."""
        ws = wb.create_sheet("Document_Metadata", 0)

        # Write metadata as key-value pairs
        data = metadata.to_worksheet_data()

        for row_idx, (key, value) in enumerate(data, start=1):
            ws.cell(row_idx, 1, key)
            ws.cell(row_idx, 2, value)

        # Style header column
        for row_idx in range(1, len(data) + 1):
            cell = ws.cell(row_idx, 1)
            if cell.value and cell.value.strip():  # Not blank row
                cell.font = Font(bold=True)

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    def _write_table_to_sheet(
        self, wb: Workbook, table: ExtractedTable, sheet_name: str
    ) -> Worksheet:
        """
        Write a single table to a worksheet.

        Args:
            wb: Workbook
            table: Table to write
            sheet_name: Name for worksheet

        Returns:
            Created worksheet
        """
        ws = wb.create_sheet(sheet_name)

        # Write data
        for row_idx, row_data in enumerate(table.data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row_idx, col_idx, cell_value)

        # Style header row
        if table.schema.has_header and table.data:
            self._apply_header_style(ws, row=1, num_cols=table.schema.column_count)

        # Apply borders
        if self.add_borders and table.data:
            self._apply_borders(ws, len(table.data), table.schema.column_count)

        # Freeze header
        if self.freeze_headers and table.schema.has_header:
            ws.freeze_panes = "A2"

        # Auto-size columns
        self._auto_size_columns(ws, table.schema.column_count)

        return ws

    def _write_logical_tables_separate(
        self, wb: Workbook, tables: List[ExtractedTable]
    ) -> None:
        """Write each logical table to a separate worksheet."""
        for idx, table in enumerate(tables, start=1):
            # Generate sheet name based on table info
            sheet_name = self._generate_sheet_name(table, idx)
            self._write_table_to_sheet(wb, table, sheet_name)

    def _write_logical_tables_single_vertical(
        self, wb: Workbook, tables: List[ExtractedTable]
    ) -> None:
        """Write all logical tables to a single worksheet, stacked vertically."""
        ws = wb.create_sheet("Logical_Tables")

        current_row = 1

        for table_idx, table in enumerate(tables):
            # Add section title if available
            if table.section_title:
                ws.cell(current_row, 1, table.section_title)
                ws.cell(current_row, 1).font = Font(bold=True, size=12)
                current_row += 1

            # Write table data
            for row_data in table.data:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(current_row, col_idx, cell_value)
                current_row += 1

            # Apply header style to first row of this table
            if table.schema.has_header:
                header_row = current_row - len(table.data)
                self._apply_header_style(
                    ws, row=header_row, num_cols=table.schema.column_count
                )

            # Add blank row separator between tables
            if table_idx < len(tables) - 1:
                current_row += 1

        # Apply borders to entire sheet
        if self.add_borders:
            self._apply_borders(ws, current_row - 1, max(t.schema.column_count for t in tables))

        # Freeze first header
        if self.freeze_headers and tables and tables[0].schema.has_header:
            ws.freeze_panes = "A2"

        # Auto-size columns
        max_cols = max(t.schema.column_count for t in tables) if tables else 0
        self._auto_size_columns(ws, max_cols)

    def _write_logical_tables_single_horizontal(
        self, wb: Workbook, tables: List[ExtractedTable]
    ) -> None:
        """Write all logical tables to a single worksheet, placed horizontally."""
        ws = wb.create_sheet("Logical_Tables")

        current_col = 1

        for table_idx, table in enumerate(tables):
            # Write table data
            for row_idx, row_data in enumerate(table.data, start=1):
                for col_offset, cell_value in enumerate(row_data):
                    ws.cell(row_idx, current_col + col_offset, cell_value)

            # Apply header style
            if table.schema.has_header:
                self._apply_header_style(
                    ws,
                    row=1,
                    num_cols=table.schema.column_count,
                    start_col=current_col,
                )

            # Apply borders to this table
            if self.add_borders:
                self._apply_borders(
                    ws,
                    len(table.data),
                    table.schema.column_count,
                    start_col=current_col,
                )

            # Move to next column position (with separator)
            current_col += table.schema.column_count + 1

        # Freeze first header
        if self.freeze_headers and tables and tables[0].schema.has_header:
            ws.freeze_panes = "A2"

        # Auto-size all columns
        self._auto_size_columns(ws, current_col - 1)

    def _write_validation_sheet(
        self, wb: Workbook, report: ValidationReport
    ) -> None:
        """Write validation report to a worksheet."""
        ws = wb.create_sheet("Validation_Report")

        # Summary section
        ws.cell(1, 1, "Validation Summary")
        ws.cell(1, 1).font = Font(bold=True, size=14)

        ws.cell(3, 1, "Overall Status")
        ws.cell(3, 2, report.overall_status.value.upper())

        status_cell = ws.cell(3, 2)
        if report.overall_status.value == "passed":
            status_cell.font = Font(color="1F7A1F", bold=True)
        elif report.overall_status.value == "warning":
            status_cell.font = Font(color="D97706", bold=True)
        else:
            status_cell.font = Font(color="9B1C1C", bold=True)

        ws.cell(4, 1, "Tables Validated")
        ws.cell(4, 2, report.tables_validated)

        ws.cell(5, 1, "Tables Passed")
        ws.cell(5, 2, report.tables_passed)

        ws.cell(6, 1, "Tables with Warnings")
        ws.cell(6, 2, report.tables_with_warnings)

        ws.cell(7, 1, "Tables Failed")
        ws.cell(7, 2, report.tables_failed)

        # Issues section
        if report.issues:
            ws.cell(9, 1, "Validation Issues")
            ws.cell(9, 1).font = Font(bold=True, size=12)

            # Headers
            headers = ["Severity", "Table", "Row", "Column", "Message"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(10, col_idx, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="0B1F3B", end_color="0B1F3B", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            # Issues
            for issue_idx, issue in enumerate(report.issues, start=11):
                ws.cell(issue_idx, 1, issue.severity.value.upper())
                ws.cell(issue_idx, 2, issue.table_name)
                ws.cell(issue_idx, 3, issue.row_index if issue.row_index is not None else "")
                ws.cell(issue_idx, 4, issue.column_name or "")
                ws.cell(issue_idx, 5, issue.message)

                # Color code severity
                severity_cell = ws.cell(issue_idx, 1)
                if issue.severity == "failed":
                    severity_cell.font = Font(color="9B1C1C", bold=True)
                elif issue.severity == "warning":
                    severity_cell.font = Font(color="D97706", bold=True)

        # Auto-size columns
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 60

    def _apply_header_style(
        self, ws: Worksheet, row: int, num_cols: int, start_col: int = 1
    ) -> None:
        """Apply header styling to a row."""
        # Get theme colors or use defaults
        if self.theme:
            header_color = self.theme.palette.primary.replace("#", "")
        else:
            header_color = "0B1F3B"  # Corporate navy blue

        for col_idx in range(start_col, start_col + num_cols):
            cell = ws.cell(row, col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=header_color, end_color=header_color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_borders(
        self,
        ws: Worksheet,
        num_rows: int,
        num_cols: int,
        start_row: int = 1,
        start_col: int = 1,
    ) -> None:
        """Apply borders to table range."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_idx in range(start_row, start_row + num_rows):
            for col_idx in range(start_col, start_col + num_cols):
                ws.cell(row_idx, col_idx).border = thin_border

    def _auto_size_columns(self, ws: Worksheet, num_cols: int) -> None:
        """Auto-size columns based on content."""
        for col_idx in range(1, num_cols + 1):
            column_letter = get_column_letter(col_idx)

            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _generate_sheet_name(
        self, table: ExtractedTable, index: int
    ) -> str:
        """
        Generate worksheet name for a table.

        Args:
            table: Table
            index: Table index

        Returns:
            Sheet name (max 31 characters for Excel)
        """
        # Try to use section title or score domain
        if table.section_title:
            name = table.section_title[:25]
        elif table.score_domain:
            name = table.score_domain.name[:25]
        else:
            name = f"Logical_Table_{index}"

        # Clean name (Excel doesn't allow certain characters)
        name = name.replace("/", "_").replace("\\", "_").replace(":", "_")
        name = name.replace("*", "_").replace("?", "_").replace("[", "_")
        name = name.replace("]", "_")

        # Truncate to 31 characters (Excel limit)
        return name[:31]
